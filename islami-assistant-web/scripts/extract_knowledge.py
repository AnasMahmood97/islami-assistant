import json
import sys
from openpyxl import load_workbook


def main(path: str):
    wb = load_workbook(path)
    ws = wb.active

    rows = []
    images = {}
    for image in getattr(ws, "_images", []):
        anchor = getattr(image, "anchor", None)
        if anchor and getattr(anchor, "_from", None):
            row_idx = anchor._from.row + 1
            images[row_idx] = f"embedded_image_row_{row_idx}"

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        question = str(row[0] or "").strip()
        answer = str(row[1] or "").strip()
        if not question or not answer:
            continue
        rows.append(
            {
                "question": question,
                "answer": answer,
                "imageRef": images.get(idx),
            }
        )

    print(json.dumps({"items": rows}, ensure_ascii=False))


if __name__ == "__main__":
    if len(sys.argv) < 2:
      raise SystemExit("Missing xlsx path argument")
    main(sys.argv[1])
